from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from openpyxl.styles import PatternFill, Alignment
from flask import Flask, request
from twilio import twiml
from timeit import default_timer as timer
import time
import Pages
# Sends the same text to mass amount
# Set num_recipients < 0 to send to all numbers in file
# Client must already be declared with tokens

# Cell colors
RED = 'ff0000'
GREEN = '00ff00'
YELLOW = 'ffff00'

class Person:
        def __init__(self, name, num):
            self.name = name
            self.number = num
        def getInfo(self):
            return self.name + ": " + self.number


# Prepends country code if not present
def formatNumber(num):
    if num[0] != '+':
        num = '+1' + num  # U.S. numbers only
    return num

# Function takes in a message to send and a name string
# Replaces 'name' in original message with first name only
def addFirstNameToMsg(message, name):
    try:
        full_name = name.split()
        name_len = len(full_name)
    except:
        print("Unable to split this name: " + name)
    else:
        if name_len >= 2:
            first_name = full_name[0]
        elif name_len is 1:
            first_name = name
    
    new_msg = message.replace("'name'", first_name)
    return new_msg

# Returns list of properly formatted phone numbers from excel sheet

def populateNumberList(sheet):
    max_row = sheet.max_row
    contactList = []
    for i in range(1, max_row + 1):
        numCell = sheet.cell(row=i, column=1)
        nameCell = sheet.cell(row=i, column=2)
        person = Person(str(nameCell.value), formatNumber(str(numCell.value)))
        contactList.append(person)
        #print("Person added to contact list: " + person.getInfo())
    return contactList


# Tracks used numbers in excel sheet, marks yellow if sent
def markSentExcel(num, sheet, color):
    count = 1
    for row in range(1, sheet.max_row + 1):
        cellObj = sheet.cell(row=row, column=1)  # Phone number
        cellColor = sheet.cell(row=row, column=3)  # Status color
        if str(cellObj.value) == num.replace("+1", ""):
            cellColor.fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")
            if color == GREEN:
                cellColor.value = "SENT"
            else:
                cellColor.value = "FAIL"
            cellColor.alignment = Alignment(horizontal='center')
            break
        count += 1

# Send same message to entire list
# msg - message to mass send; client - twilio client; numbers - list of numbers as strings
# twil_num - twilio number to send msg from; sheet - excel sheet for marking sent


def massSendSMS(msg, numbers, sheet, self):
    start = timer()
    count = 1
    output = None
    for num in numbers:
        try:
            message = addFirstNameToMsg(msg, num.name)
            #print("Message to " + str(num.number) + " " + message)
            self.client.messages.create(body=message, from_=self.twil_num, to=num.number)
        except:
            markSentExcel(num.number, sheet, RED)
            output = "Message send failure to " + num.number
        else:
            markSentExcel(num.number, sheet, GREEN)
            output = "Message sent to " + num.number
        if self.surpress_ckbx.isChecked() == False:
            self.output_textbox.appendPlainText(output)
        self.progress.setValue((count/len(numbers)) * 100)
        QApplication.processEvents()
        count += 1
        time.sleep(0.3)
    end = timer()
    return end - start
